#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# The MIT License (MIT)
# 
# Copyright (c) 2022, Roland Rickborn (r_2@gmx.net)
# 
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
# 
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
# 
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
# THE SOFTWARE.
#
# ---------------------------------------------------------------------------

import csv
import math
import sys
from datetime import datetime

import openpyxl

import enums
import utils


def write_init_output_file(outputFilename='', outputColumns={}):
    outputFilename = utils.get_save_filename(outputFilename)
    retval = outputFilename
    with open(outputFilename, 'w', newline='', encoding='utf-8') as csvfile:
        csvwriter = csv.writer(csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        _l = list(outputColumns.values())
        _l[0] = u'\uFEFF' + _l[0]
        csvwriter.writerow(_l)
    return retval

def append_to_output_file(outputfile='', data={}):
    with open(outputfile, 'a', newline='', encoding='utf-8') as csvfile:
        csvwriter = csv.writer(csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        new_line = []
        for item in list(data.values()):
            new_line.append(str(item).replace('\n',''))
        csvwriter.writerow(new_line)

def remove_duplicate_keywords(input=''):
    retval = ''
    unique_keywords = []
    for item in input.split(';'):
        if not item.lower() in unique_keywords:
            unique_keywords.append(item.lower())
    retval = ';'.join(unique_keywords)
    return retval

def convert_excel_to_csv(filename=''):
    retval = []
    wb = openpyxl.load_workbook('{}.xlsx'.format(filename))
    ws = wb.active
    end_of_ws_reached = False
    counter = 2
    limit = 3000
    data_to_be_appended = []
    titles = []
    while not end_of_ws_reached:
        _v = ws['A{}'.format(counter)].value
        if _v == None:
            end_of_ws_reached = True
        else:
            if not ws['A{}'.format(counter)].value == None and ws['A{}'.format(counter)].value not in titles and len(ws['A{}'.format(counter)].value) < 60:
                titles.append(ws['A{}'.format(counter)].value)
                my_title = ws['A{}'.format(counter)].value
            elif ws['A{}'.format(counter)].value == None:
                break
            elif ws['A{}'.format(counter)].value in titles:
                print('Title "{}" already exists ({})'.format(ws['A{}'.format(counter)].value, counter))
                sys.exit()
            elif len(ws['A{}'.format(counter)].value) >= 60:
                my_title = '{}...'.format(ws['A{}'.format(counter)].value[0:57])
            if not ws['B{}'.format(counter)].value == None:
                my_url = ws['B{}'.format(counter)].value
            else:
                my_url = ''
            if not ws['C{}'.format(counter)].value == None:
                my_keywords = remove_duplicate_keywords(ws['C{}'.format(counter)].value)
            else:
                my_keywords = ''
            if not ws['D{}'.format(counter)].value == None:
                my_match_similar_keywords = ws['D{}'.format(counter)].value
            else:
                my_match_similar_keywords = ''
            if not ws['E{}'.format(counter)].value == None:
                my_state = ws['E{}'.format(counter)].value
            else:
                my_state = ''
            if not ws['F{}'.format(counter)].value == None:
                my_description = ws['F{}'.format(counter)].value
            else:
                my_description = ''
            if not ws['G{}'.format(counter)].value == None:
                my_reserved_keywords = ws['G{}'.format(counter)].value
            else:
                my_reserved_keywords = ''
            if not ws['H{}'.format(counter)].value == None:
                my_categories = ws['H{}'.format(counter)].value
            else:
                my_categories = ''
            if not ws['I{}'.format(counter)].value == None:
                try:
                    _d = datetime.strptime(ws['I{}'.format(counter)].value,'%Y-%m-%dT%H:%M:%S+00') #Date, 2022-11-27T07:00:00+00
                    my_start_date = _d.strftime('%Y-%m-%dT%H:%M:%S+00')
                except:
                    try:
                        my_start_date = ws['I{}'.format(counter)].value.strftime('%Y-%m-%dT%H:%M:%S+00')
                    except:
                        my_start_date = ws['I{}'.format(counter)]
            else:
                my_start_date = ''
            if not ws['J{}'.format(counter)].value == None:
                try:
                    _d = datetime.strptime(ws['J{}'.format(counter)].value,'%Y-%m-%dT%H:%M:%S+00') #Date, 2022-11-28T19:00:00+00
                    my_end_date = _d.strftime('%Y-%m-%dT%H:%M:%S+00')
                except:
                    try:
                        my_end_date = ws['J{}'.format(counter)].value.strftime('%Y-%m-%dT%H:%M:%S+00')
                    except:
                        my_end_date = ws['J{}'.format(counter)].value
            else:
                my_end_date = ''
            if not ws['K{}'.format(counter)].value == None:
                my_country_region = ws['K{}'.format(counter)].value
            else:
                my_country_region = ''
            if not ws['L{}'.format(counter)].value == None:
                my_use_AAD_location = ws['L{}'.format(counter)].value
            else:
                my_use_AAD_location = ''
            if not ws['M{}'.format(counter)].value == None:
                my_groups = ws['M{}'.format(counter)].value
            else:
                my_groups = ''
            if not ws['N{}'.format(counter)].value == None:
                my_device_and_OS = ws['N{}'.format(counter)].value
            else:
                my_device_and_OS = ''
            if not ws['O{}'.format(counter)].value == None:
                my_targeted_variations = ws['O{}'.format(counter)].value
            else:
                my_targeted_variations = ''
            if not ws['P{}'.format(counter)].value == None:
                try:
                    _d = datetime.strptime(ws['P{}'.format(counter)].value,'%m/%d/%Y') #Date, 11/24/2022
                    my_start_date = _d.strftime('%m/%d/%Y')
                except:
                    try:
                        my_last_modified = ws['P{}'.format(counter)].value.strftime('%m/%d/%Y')
                    except:
                        my_last_modified = ws['P{}'.format(counter)]
            else:
                my_last_modified = ''
            if not ws['Q{}'.format(counter)].value == None:
                my_last_modified_by = ws['Q{}'.format(counter)].value
            else:
                my_last_modified_by = ''
            if not ws['R{}'.format(counter)].value == None:
                my_id = ws['R{}'.format(counter)].value
            else:
                my_id = ''
            my_data = {
                'A': my_title,
                'B': my_url,
                'C': my_keywords,
                'D': my_match_similar_keywords,
                'E': my_state,
                'F': my_description,
                'G': my_reserved_keywords,
                'H': my_categories,
                'I': my_start_date,
                'J': my_end_date,
                'K': my_country_region,
                'L': my_use_AAD_location,
                'M': my_groups,
                'N': my_device_and_OS,
                'O': my_targeted_variations,
                'P': my_last_modified,
                'Q': my_last_modified_by,
                'R': my_id
            }
            data_to_be_appended.append(my_data)
            counter += 1
    counter = 0
    number_of_files = math.ceil(len(data_to_be_appended)/limit)
    if number_of_files > 1:
        for i in range(0,number_of_files):
            my_output_filename = write_init_output_file('{}_{}.csv'.format(filename,i+1), my_output_columns)
            print('Output file: {}'.format(my_output_filename))
            retval.append(my_output_filename)
            for j in range(0,limit):
                if counter < len(data_to_be_appended):
                    append_to_output_file(my_output_filename, data_to_be_appended[counter])
                    counter += 1
    else:
        my_output_filename = write_init_output_file('{}.csv'.format(filename), my_output_columns)
        print('Output file: {}'.format(my_output_filename))
        retval.append(my_output_filename)
        for j in range(0,limit):
            if counter < len(data_to_be_appended):
                append_to_output_file(my_output_filename, data_to_be_appended[counter])
                counter += 1
    return retval

my_output_columns = enums.Enums.columns

filename = ''
if __name__ == "__main__":
    if len(sys.argv) > 1:
        filename = sys.argv[1].split('.xlsx')[0]
        output = convert_excel_to_csv(filename)
