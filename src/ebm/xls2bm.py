#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# The MIT License (MIT)
#
# Copyright (c) 2023, Roland Rickborn (r_2@gmx.net)
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
# THE SOFTWARE.
#
# ---------------------------------------------------------------------------

import csv
import math
import sys

import openpyxl

import src.ebm.bookmark as bookmark
import src.ebm.bookmark_shelf as bookmark_shelf
import src.ebm.utils as utils


def write_init_output_file(outputFilename: str, outputColumns: list):
    outputFilename = utils.get_save_filename(outputFilename)
    retval = outputFilename
    with open(outputFilename, 'w', newline='', encoding='utf-8') as csvfile:
        csvwriter = csv.writer(
            csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        outputColumns[0] = u'\uFEFF' + outputColumns[0]
        csvwriter.writerow(outputColumns)
    return retval


def append_to_output_file(outputfile: str, data: list):
    with open(outputfile, 'a', newline='', encoding='utf-8') as csvfile:
        csvwriter = csv.writer(
            csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
        new_line = []
        for item in data:
            new_line.append(str(item).replace('\n', ''))
        csvwriter.writerow(new_line)


def read_input_file(filename: str):
    retval = bookmark_shelf.BookmarkShelf()
    wb = openpyxl.load_workbook('{}.xlsx'.format(filename))
    ws = wb.active
    for r in range(2, ws.max_row):
        my_bookmark = bookmark.Bookmark(
            title=ws['{}{}'.format('A', r)].value,
            url=ws['{}{}'.format('B', r)].value,
            keywords=ws['{}{}'.format('C', r)].value,
            match_similar_keywords=ws['{}{}'.format('D', r)].value,
            state=ws['{}{}'.format('E', r)].value,
            description=ws['{}{}'.format('F', r)].value,
            reserved_keywords=ws['{}{}'.format('G', r)].value,
            categories=ws['{}{}'.format('H', r)].value,
            start_date=ws['{}{}'.format('I', r)].value,
            end_date=ws['{}{}'.format('J', r)].value,
            country_region=ws['{}{}'.format('K', r)].value,
            use_aad_location=ws['{}{}'.format('L', r)].value,
            groups=ws['{}{}'.format('M', r)].value,
            device_and_os=ws['{}{}'.format('N', r)].value,
            targeted_variations=ws['{}{}'.format('O', r)].value,
            last_modified=ws['{}{}'.format('P', r)].value,
            last_modified_by=ws['{}{}'.format('Q', r)].value,
            id=ws['{}{}'.format('R', r)].value
        )
        retval.add_bookmark(my_bookmark)
    return retval


def convert_excel_to_csv(filename: str):
    retval = []
    my_output_columns = list(bookmark.Bookmark.get_columns().values())
    my_shelf = read_input_file(filename)
    limit = 3000
    counter = 0
    number_of_files = math.ceil(len(my_shelf.get_bookmarks())/limit)
    if number_of_files > 1:
        for i in range(0, number_of_files):
            my_output_filename = write_init_output_file('{}_{}.csv'.format(
                filename, i+1), my_output_columns)
            retval.append(my_output_filename)
            for j in range(0, limit):
                my_keys = list(my_shelf.get_bookmarks())
                if counter < len(my_keys):
                    my_bm = my_shelf.get_bookmark(my_keys[counter])
                    append_to_output_file(
                        my_output_filename, my_bm.to_string())
                    counter += 1
    else:
        my_output_filename = write_init_output_file('{}.csv'.format(
            filename), my_output_columns)
        retval.append(my_output_filename)
        for j in range(0, limit):
            my_keys = list(my_shelf.get_bookmarks())
            if counter < len(my_keys):
                my_bm = my_shelf.get_bookmark(my_keys[counter])
                append_to_output_file(my_output_filename, my_bm.to_string())
                counter += 1
    return ', '.join(retval)


filename = ''
if __name__ == "__main__":
    if len(sys.argv) > 1:
        filename = sys.argv[1].split('.xlsx')[0]
        output = convert_excel_to_csv(filename)
        print('Output file: {}'.format(output))
