#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# The MIT License (MIT)
#
# Copyright (c) 2023, Roland Rickborn (r_2@gmx.net)
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
# THE SOFTWARE.
#
# ---------------------------------------------------------------------------

import csv
import datetime as dt
import locale

import openpyxl

import src.ebm.bookmark as bookmark
import src.ebm.utils as utils


def read_input_file(input_filename: str):
    retval = {}
    my_output_columns = bookmark.Bookmark.get_columns()
    header_row_keys = list(my_output_columns.keys())
    header_checked = False
    with open(input_filename, newline='', encoding='utf-8') as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        for row in csv_reader:
            if not header_checked:
                if validate_header(row):
                    header_checked = True
                else:
                    raise ValidationError(
                        'Header of CSV file not correct')
            t = {}
            id = row[-1]
            for col in range(0, len(row)):
                t[header_row_keys[col]] = row[col]
            retval[id] = t
    return retval


def validate_header(header_row: list):
    my_output_columns = bookmark.Bookmark.get_columns()
    header_row_values = list(my_output_columns.values())
    if header_row_values == header_row:
        return True
    else:
        return False


def get_date_by_str(datetimestr: str):
    retval = ''
    if 'T' in datetimestr and '+' in datetimestr:
        try:
            d = dt.datetime.strptime(datetimestr+'00', '%Y-%m-%dT%H:%M:%S%z')
            retval = d.replace(tzinfo=None)
        except Exception:
            retval = datetimestr
    else:
        try:
            d = dt.datetime.strptime(datetimestr, '%m/%d/%Y')
            retval = d.replace(tzinfo=None)
        except Exception:
            retval = datetimestr
    return retval


def get_date_format_by_str(datetimestr: str):
    retval = ''
    loc = locale.getlocale()
    if 'T' in datetimestr and '+' in datetimestr:
        if loc[0].startswith('de'):
            retval = 'dd.mm.yyyy HH:MM:SS'
        else:
            retval = 'mm/dd/yyyy HH:MM:SS'
    elif '/' in datetimestr:
        if loc[0].startswith('de'):
            retval = 'dd.mm.yyyy'
        else:
            retval = 'mm/dd/yyyy'
    return retval


def convert_csv_to_excel(filename: str):
    my_input_data = read_input_file('{}.csv'.format(filename))
    wb = openpyxl.Workbook()
    wb.iso_dates = True
    ws = wb.active
    ws.title = filename.split('\\')[-1]
    my_input_data_keys = list(my_input_data.keys())
    for item in range(0, len(my_input_data_keys)):
        bookmark_id = my_input_data_keys[item]
        cell_number = item+1
        cell_chars = list(my_input_data[bookmark_id].keys())
        for cell_char in cell_chars:
            if cell_number == 1:  # Write header line
                ws['{}{}'.format(cell_char, cell_number)] = my_input_data[
                    bookmark_id][cell_char]
                ws['{}{}'.format(
                    cell_char, cell_number)].font = openpyxl.styles.Font(
                        bold=True)
            elif cell_char in ['I', 'J', 'P']:  # Detect datetime objects
                print('HELLO')
                my_date = get_date_by_str(
                    my_input_data[bookmark_id][cell_char])
                ws['{}{}'.format(cell_char, cell_number)] = my_date
                if isinstance(my_date, dt.datetime):
                    my_date_format = get_date_format_by_str(
                        my_input_data[bookmark_id][cell_char])
                    ws['{}{}'.format(
                        cell_char, cell_number)].number_format = my_date_format
            else:
                ws['{}{}'.format(
                    cell_char, cell_number)] = my_input_data[bookmark_id][
                        cell_char]
    ws.auto_filter.ref = ws.dimensions
    new_filename = utils.get_save_filename('{}.xlsx'.format(filename))
    wb.save(new_filename)
    return new_filename


class ValidationError(Exception):
    pass
