#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# The MIT License (MIT)
# 
# Copyright (c) 2022, Roland Rickborn (r_2@gmx.net)
# 
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
# 
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
# 
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
# THE SOFTWARE.
#
# ---------------------------------------------------------------------------

import csv
import sys
from datetime import datetime

import openpyxl

import enums
import utils


def read_input_file(inputFilename=''):
    global my_output_columns
    retval = {}
    header_row_keys = list(my_output_columns.keys())
    header_row_values = list(my_output_columns.values())
    with open(inputFilename, newline='', encoding='utf-8') as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=',')
        line_count = 0
        for row in csv_reader:
            if row == 0:
                for col in range(0,len(row)):
                    if not row[col].endswith(header_row_values[col]):
                        print('Error')
                        break
            else:
                t = {}
                id = row[-1]
                for col in range(0,len(row)):
                    t[header_row_keys[col]] = row[col]
                retval[id] = t
            line_count += 1
    return retval

def convert_csv_to_excel(filename=''):
    my_input_data = read_input_file('{}.csv'.format(filename))
    wb = openpyxl.Workbook()
    wb.iso_dates = True
    ws = wb.active
    ws.title = filename
    my_input_data_keys = list(my_input_data.keys())
    for item in range(0,len(my_input_data_keys)):
        bookmark_id = my_input_data_keys[item]
        cell_number = item+1
        cell_chars = list(my_input_data[bookmark_id].keys())
        for cell_char in cell_chars:
            if cell_number == 1:
                ws['{}{}'.format(cell_char, cell_number)] = my_input_data[bookmark_id][cell_char]
                ws['{}{}'.format(cell_char, cell_number)].font = openpyxl.styles.Font(bold = True)
            elif cell_char == 'P':
                try:
                    d = datetime.strptime(my_input_data[bookmark_id][cell_char], '%m/%d/%Y')
                    ws['{}{}'.format(cell_char, cell_number)] = d
                    ws['{}{}'.format(cell_char, cell_number)].number_format = 'mm.dd.yyyy'
                except:
                    ws['{}{}'.format(cell_char, cell_number)] = my_input_data[bookmark_id][cell_char]
            else:
                ws['{}{}'.format(cell_char, cell_number)] = my_input_data[bookmark_id][cell_char]
    ws.auto_filter.ref = ws.dimensions
    new_filename = utils.get_save_filename('{}.xlsx'.format(filename))
    wb.save(new_filename)
    print('Output file: {}'.format(new_filename))
    return new_filename

my_output_columns = enums.Enums.columns

filename = ''
if __name__ == "__main__":
    if len(sys.argv) > 1:
        filename = sys.argv[1].split('.csv')[0]
        output = convert_csv_to_excel(filename)
