#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# The MIT License (MIT)
#
# Copyright (c) 2025, Roland Rickborn (r_2@gmx.net)
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in
# all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
# THE SOFTWARE.
#
# ---------------------------------------------------------------------------

import datetime as dt
import locale

import openpyxl
import pytest

import src.ebm.bm2xls as bm2xls
import tests.ebm_fixtures as fix


@pytest.fixture(scope='session')
def input_filename(tmpdir_factory):
    fn = tmpdir_factory.mktemp('data').join('{}.csv'.format(fix.FILENAME))
    f = open(fn, 'w')
    f.write(fix.CSV_FILE_GOOD)
    f.close()
    return str(fn)


class TestBm2xlxReadInput(object):

    def test_read_input(self, input_filename):
        retval = bm2xls.read_input_file(input_filename)
        assert fix.ID_GOOD in retval
        assert retval[fix.ID_GOOD]['A'] == fix.TITLE_GOOD
        assert retval[fix.ID_GOOD]['B'] == fix.URL_GOOD
        assert retval[fix.ID_GOOD]['C'] == fix.KEYWORDS_GOOD
        assert retval[fix.ID_GOOD]['D'] == fix.MATCH_SIMILAR_KEYWORDS_GOOD
        assert retval[fix.ID_GOOD]['E'] == fix.STATE_GOOD
        assert retval[fix.ID_GOOD]['F'] == fix.DESCRIPTION
        assert retval[fix.ID_GOOD]['J'] == fix.END_DATE
        assert retval[fix.ID_GOOD]['L'] == fix.USE_AAD_LOCATION_FALSE_GOOD
        assert retval[fix.ID_GOOD]['P'] == fix.LAST_MODIFIED
        assert retval[fix.ID_GOOD]['Q'] == fix.LAST_MODIFIED_BY

    def test_validate_header_good(self):
        header_row = fix.HEADER_GOOD.split(',')
        retval = bm2xls.validate_header(header_row)
        assert retval is True

    def test_validate_header_bad(self):
        header_row = fix.HEADER_BAD.split(',')
        retval = bm2xls.validate_header(header_row)
        assert retval is False


class TestBm2xlxWriteXlsx(object):

    def test_convert_csv_to_excel(self, input_filename):
        _use_aad_location = fix.USE_AAD_LOCATION_FALSE_GOOD
        filename = input_filename.split('.csv')[0]
        output = bm2xls.convert_csv_to_excel(filename)
        assert output.endswith('{}.xlsx'.format(fix.FILENAME))
        wb = openpyxl.load_workbook(output)
        ws = wb.active
        assert ws['{}{}'.format('A', 2)].value == fix.TITLE_GOOD
        assert ws['{}{}'.format('B', 2)].value == fix.URL_GOOD
        assert ws['{}{}'.format('C', 2)].value == fix.KEYWORDS_GOOD
        _ms = fix.MATCH_SIMILAR_KEYWORDS_GOOD
        assert ws['{}{}'.format('D', 2)].value == _ms
        assert ws['{}{}'.format('E', 2)].value == fix.STATE_GOOD
        assert ws['{}{}'.format('F', 2)].value == fix.DESCRIPTION
        assert ws['{}{}'.format('L', 2)].value == _use_aad_location
        my_last_modified = ws['{}{}'.format('P', 2)].value
        assert my_last_modified.strftime('%m/%d/%Y') == fix.LAST_MODIFIED
        assert ws['{}{}'.format('Q', 2)].value == fix.LAST_MODIFIED_BY


class TestBm2xlxDates(object):

    def test_last_modified_date(self):
        my_date = bm2xls.get_date_by_str(fix.LAST_MODIFIED)
        assert my_date == dt.datetime(2022, 12, 28, 0, 0)

    def test_start_end_date(self):
        my_date = bm2xls.get_date_by_str(fix.END_DATE)
        assert my_date == dt.datetime(2022, 12, 29, 7, 30)

    def test_date_string(self):
        my_date = bm2xls.get_date_by_str('29-12-2022')
        assert my_date == '29-12-2022'

    def test_false_date_string(self):
        my_date = bm2xls.get_date_by_str('29-12-2022T07:30:00+00')
        assert my_date == '29-12-2022T07:30:00+00'


class TestBm2xlxFormats(object):

    def test_last_modified_format_de(self):
        locale.setlocale(locale.LC_ALL, 'de_DE')
        my_format = bm2xls.get_date_format_by_str(fix.LAST_MODIFIED)
        locale.setlocale(locale.LC_ALL, '')
        assert my_format == 'dd.mm.yyyy'

    def _last_modified_format_en(self):
        locale.setlocale(locale.LC_ALL, 'en_US')
        my_format = bm2xls.get_date_format_by_str(fix.LAST_MODIFIED)
        locale.setlocale(locale.LC_ALL, '')
        assert my_format == 'mm/dd/yyyy'

    def test_start_end_format_de(self):
        locale.setlocale(locale.LC_ALL, 'de_AT')
        my_format = bm2xls.get_date_format_by_str(fix.END_DATE)
        locale.setlocale(locale.LC_ALL, '')
        assert my_format == 'dd.mm.yyyy HH:MM:SS'

    def test_start_end_format_en(self):
        locale.setlocale(locale.LC_ALL, 'en_GB')
        my_format = bm2xls.get_date_format_by_str(fix.END_DATE)
        locale.setlocale(locale.LC_ALL, '')
        assert my_format == 'mm/dd/yyyy HH:MM:SS'

    def test_not_date_format(self):
        my_format = bm2xls.get_date_format_by_str('anything')
        assert my_format == ''
